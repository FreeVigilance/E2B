import csv
import pathlib

import openpyxl
from django.core.management import BaseCommand

STRENGTH_DOSE_UNITS = {
    'amount of substance',
    'amount of substance (dissolved particles)',
    'Substance Units',
    'Areic Substance Units',
    'Substance Content Units',
    'Substance Ratio Or Substance Fraction Units',
    'Substance Concentration Units',

    'mass',
    'SI Mass Units',
    'Molar Mass Units',
    'Areic Mass Units',
    'Lineic Mass Units',
    'Mass Concentration Units',
    'Mass Ratio Or Mass Fraction Or Mass Content Units',
    'Mass Concentration Units',

    'volume',
    'SI Volume Units',
    'Volume Fraction Units',
    'Volume Content Units',

    'radioactivity',

    'arbitrary',
    'Arbitrary Concentration Units',
    'Arbitrary Concentration Content Units',

    'fraction',
    'Number Concentration Units'
}

TIME_UNITS = {
    'time',
    'Time Units'
}


def get_relevant_codes_from_phin_vads_file(file_path: pathlib.Path) -> dict:
    with file_path.open() as f:
        reader = csv.reader(f, delimiter='\t')

        # Skip the first four rows
        for _ in range(4):
            next(reader)

        codes = dict()
        for data in reader:
            code, _, preferred_name = data[:3]
            property = preferred_name.split('[')[-1].rstrip(']') if ']' in preferred_name else None
            codes[code] = property
    return codes


def get_relevant_codes_from_ucum_examples_file(file_path: pathlib.Path) -> dict:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = next(worksheet for worksheet in workbook.worksheets if worksheet.title.startswith("Example UCUM"))

    codes = dict()
    for code, name in worksheet.iter_rows(min_row=2, min_col=2, max_col=3, max_row=worksheet.max_row):
        codes[code.value] = name.value.strip()

    return codes


class Command(BaseCommand):
    help = 'Prepare UCUM file for upload_ucum command'

    def add_arguments(self, parser):
        parser.add_argument('phin_vads', type=pathlib.Path)
        parser.add_argument('ucum_examples', type=pathlib.Path)
        parser.add_argument('output', type=pathlib.Path)

    def handle(self, *args, **options):
        phin_vads_file = pathlib.Path(options['phin_vads'])
        ucum_examples_file = pathlib.Path(options['ucum_examples'])

        with open(options['output'], 'w') as f:
            writer = csv.writer(f)
            code_to_name = get_relevant_codes_from_ucum_examples_file(ucum_examples_file)
            code_to_property = get_relevant_codes_from_phin_vads_file(phin_vads_file)
            for code in sorted(code_to_property.keys() & code_to_name.keys()):
                property = code_to_property[code]
                if property in TIME_UNITS:
                    writer.writerow([code, code_to_name[code], 'time'])
                elif property in STRENGTH_DOSE_UNITS:
                    writer.writerow([code, code_to_name[code], 'strength_dose'])
                else:
                    writer.writerow([code, code_to_name[code], 'other'])
